"""Build a formula-driven Malaysia monthly payroll workbook.

Generates ``emp_payroll_2606.xlsx`` from the master employee list and the
June 2026 input image (``payroll_2606.png``).  All statutory contributions
(EPF/KWSP, SOCSO/PERKESO, EIS and PCB) are derived with Excel formulas that
reference an auditable ``Rates`` sheet, so the file recalculates whenever the
monthly inputs change.

Statutory basis (verified against ``emp_payroll_2605.xlsx`` May figures):
  * EPF   : employee 11%; employer 13% (wage <= RM5,000) / 12% (> RM5,000).
  * SOCSO : PERKESO First/Second Category table, RM100 bands, ceiling RM6,000.
            Amount = round-half-up-to-5-sen of (rate x band midpoint).
  * EIS   : 0.2% each side, same band table, ceiling RM6,000.
  * PCB   : MTD estimate (annualised, LHDN resident brackets YA2024/25) with a
            manual override column that always wins.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "emp_payroll_2606.xlsx"
FONT = "Arial"

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------
BLUE = Font(name=FONT, size=10, color="0000FF")          # user inputs
BLACK = Font(name=FONT, size=10, color="000000")         # formulas
GREEN = Font(name=FONT, size=10, color="008000")         # cross-sheet links
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HDR2 = Font(name=FONT, size=10, bold=True, color="000000")
TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

FILL_HDR = PatternFill("solid", fgColor="1F3864")        # dark blue header
FILL_HDR_IN = PatternFill("solid", fgColor="2E75B6")     # input group header
FILL_HDR_CALC = PatternFill("solid", fgColor="548235")   # calc group header
FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")       # light blue input bg
FILL_VERIFY = PatternFill("solid", fgColor="FFFF00")      # yellow = verify
FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")
FILL_PARAM = PatternFill("solid", fgColor="FFF2CC")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

MONEY = '#,##0.00;(#,##0.00);"-"'
MONEY0 = '#,##0;(#,##0);"-"'
PCT = '0.00%'

# ===========================================================================
# 1. RATES / STATUTORY TABLES SHEET  (built first so Payroll can reference it)
# ===========================================================================
wb = Workbook()
rates = wb.active
rates.title = "Rates"
addr: dict[str, str] = {}   # symbolic name -> absolute address on Rates sheet


def rcell(row, col, value, *, font=BLACK, fmt=None, fill=None,
          align=None, border=True, name=None):
    c = rates.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    c.alignment = align or LEFT
    if border:
        c.border = BORDER
    if name:
        addr[name] = f"Rates!${get_column_letter(col)}${row}"
    return c


r = 1
rcell(r, 1, "STATUTORY RATES & CONTRIBUTION TABLES — MALAYSIA", font=TITLE, border=False)
r += 1
rcell(r, 1, "Effective rates: EPF (KWSP), SOCSO/EIS (PERKESO) ceiling RM6,000 from 1 Oct 2024, "
            "PCB brackets YA2024/2025. Edit the yellow parameter cells if rates change.",
      font=NOTE, border=False)
r += 2

# --- Parameters block ------------------------------------------------------
rcell(r, 1, "PARAMETER", font=HDR2, fill=FILL_HDR_CALC)
rcell(r, 2, "VALUE", font=HDR2, fill=FILL_HDR_CALC, align=CTR)
rcell(r, 3, "NOTES", font=HDR2, fill=FILL_HDR_CALC)
rates.cell(r, 2).font = HDR
rates.cell(r, 1).font = HDR
rates.cell(r, 3).font = HDR
r += 1
params = [
    ("EPF employee rate", 0.11, PCT, "epf_emp", "Malaysian/PR under 60"),
    ("EPF employer rate (wage <= threshold)", 0.13, PCT, "epf_er_lo", "13% for wages up to RM5,000"),
    ("EPF employer rate (wage > threshold)", 0.12, PCT, "epf_er_hi", "12% for wages above RM5,000"),
    ("EPF employer threshold (RM)", 5000, MONEY0, "epf_thresh", "Total monthly remuneration test"),
    ("SOCSO / EIS wage ceiling (RM)", 6000, MONEY0, "ceiling", "Contribution capped at this wage"),
    ("SOCSO Cat 1 employee %", 0.005, "0.000%", "soc1_emp", "Injury + Invalidity (age < 60)"),
    ("SOCSO Cat 1 employer %", 0.0175, "0.000%", "soc1_er", "Injury + Invalidity (age < 60)"),
    ("SOCSO Cat 2 employer %", 0.0125, "0.000%", "soc2_er", "Injury only (age >= 60)"),
    ("EIS rate (each side) %", 0.002, "0.000%", "eis_rate", "Employer & employee, age 18-60"),
    ("PCB personal relief (RM/yr)", 9000, MONEY0, "relief_personal", "Self"),
    ("PCB EPF+life relief cap (RM/yr)", 4000, MONEY0, "relief_epf", "EPF portion of the RM7,000 cap"),
    ("PCB rebate if chargeable <= RM35,000", 400, MONEY0, "rebate", "Section 6A individual rebate"),
    ("Default working days / month", 26, "0", "workdays", "Used for unpaid-leave proration"),
]
for label, val, fmt, key, note in params:
    rcell(r, 1, label)
    rcell(r, 2, val, font=BLUE, fmt=fmt, fill=FILL_PARAM, align=RIGHT, name=key)
    rcell(r, 3, note, font=NOTE)
    r += 1

r += 1
# --- Income-tax bracket table (YA2024/2025 resident) -----------------------
rcell(r, 1, "RESIDENT INCOME-TAX BRACKETS (YA2024/2025) — used for PCB estimate",
      font=HDR2, fill=FILL_HDR_CALC)
rates.cell(r, 1).font = HDR
for cc in (2, 3):
    rcell(r, cc, "", fill=FILL_HDR_CALC)
r += 1
rcell(r, 1, "Chargeable from (RM)", font=HDR2, fill=FILL_TOTAL, align=CTR)
rcell(r, 2, "Marginal rate", font=HDR2, fill=FILL_TOTAL, align=CTR)
rcell(r, 3, "Cumulative tax at 'from' (RM)", font=HDR2, fill=FILL_TOTAL, align=CTR)
r += 1
tax_start = r
tax_rows = [
    (0, 0.00, 0),
    (5000, 0.01, 0),
    (20000, 0.03, 150),
    (35000, 0.06, 600),
    (50000, 0.11, 1500),
    (70000, 0.19, 3700),
    (100000, 0.25, 9400),
    (400000, 0.26, 84400),
    (600000, 0.28, 136400),
    (2000000, 0.30, 528400),
]
for frm, rate, cum in tax_rows:
    rcell(r, 1, frm, fmt=MONEY0, align=RIGHT)
    rcell(r, 2, rate, fmt="0%", align=RIGHT)
    rcell(r, 3, cum, fmt=MONEY0, align=RIGHT)
    r += 1
tax_end = r - 1
addr["tax_tbl"] = f"Rates!$A${tax_start}:$C${tax_end}"

r += 1
# --- SOCSO / EIS band table (auto-generated, formula-driven) ---------------
rcell(r, 1, "SOCSO (PERKESO) + EIS CONTRIBUTION TABLE — RM100 bands, ceiling RM6,000",
      font=HDR2, fill=FILL_HDR_CALC)
rates.cell(r, 1).font = HDR
for cc in range(2, 10):
    rcell(r, cc, "", fill=FILL_HDR_CALC)
r += 1
soc_hdr = ["Key (>=)", "Wage band (RM)", "Lower", "Upper", "Assumed wage",
           "SOCSO C1 employee", "SOCSO C1 employer", "SOCSO C2 employer", "EIS (each)"]
for i, h in enumerate(soc_hdr, start=1):
    rcell(r, i, h, font=HDR2, fill=FILL_TOTAL, align=CTR)
r += 1
soc_start = r

# Band edges: fine bands at the low end, then RM100 steps up to 6,000.
edges = [0, 30, 50, 70] + list(range(100, 6001, 100))
bands = [(edges[i], edges[i + 1]) for i in range(len(edges) - 1)]
bands.append((6000, None))   # ceiling / cap row

for lower, upper in bands:
    key = 0 if lower == 0 else round(lower + 0.01, 2)
    if upper is None:                       # cap row: wages above ceiling
        label = "6,000.01 and above"
        mid_formula = f"={addr['ceiling']}-50"   # 6000-50 = 5950 midpoint
        lower_disp, upper_disp = 6000.01, ""
    else:
        label = f"{lower + 0.01:,.2f} - {upper:,.2f}"
        mid_formula = f"=(C{r}+D{r})/2"
        lower_disp, upper_disp = lower, upper
    rcell(r, 1, key, fmt="#,##0.00", align=RIGHT)
    rcell(r, 2, label, align=LEFT)
    rcell(r, 3, lower_disp if lower_disp != "" else "", fmt=MONEY0, align=RIGHT)
    rcell(r, 4, upper_disp if upper_disp != "" else "", fmt=MONEY0, align=RIGHT)
    rcell(r, 5, mid_formula, fmt=MONEY0, align=RIGHT)
    # contribution = round-half-up to 5 sen of (rate x assumed wage)
    rcell(r, 6, f"=MROUND({addr['soc1_emp']}*E{r},0.05)", fmt=MONEY, align=RIGHT)
    rcell(r, 7, f"=MROUND({addr['soc1_er']}*E{r},0.05)", fmt=MONEY, align=RIGHT)
    rcell(r, 8, f"=MROUND({addr['soc2_er']}*E{r},0.05)", fmt=MONEY, align=RIGHT)
    rcell(r, 9, f"=MROUND({addr['eis_rate']}*E{r},0.05)", fmt=MONEY, align=RIGHT)
    r += 1
soc_end = r - 1
addr["soc_tbl"] = f"Rates!$A${soc_start}:$I${soc_end}"
# VLOOKUP column indexes inside soc_tbl
SOC_C1_EMP, SOC_C1_ER, SOC_C2_ER, EIS_COL = 6, 7, 8, 9

for col, w in zip("ABCDEFGHI", (12, 22, 11, 11, 13, 15, 15, 15, 12)):
    rates.column_dimensions[col].width = w
rates.sheet_view.showGridLines = False

# ===========================================================================
# 2. PAYROLL SHEET
# ===========================================================================
pay = wb.create_sheet("Payroll", 0)   # make it the first sheet

# Column plan: (header, group, kind, width, number_format)
#   group: 'emp' | 'in' | 'calc'
#   kind : 'text' | 'input' | 'formula'
COLS = [
    ("EMP_ID",              "emp",  "text",    10, None),
    ("Name",                "emp",  "text",    22, None),
    ("Bank Name",           "emp",  "text",    13, None),
    ("Bank Acct No",        "emp",  "text",    16, None),
    ("Type",                "emp",  "input",   11, None),
    ("Age\n>= 60?",         "emp",  "input",   7,  None),
    ("Apply\nEPF?",         "emp",  "input",   7,  None),
    ("Apply\nSOCSO\n+EIS?", "emp",  "input",   8,  None),
    ("Basic",               "in",   "input",   9,  MONEY),
    ("Petrol\nAllow",       "in",   "input",   8,  MONEY),
    ("Incentive",           "in",   "input",   9,  MONEY),
    ("Dishwash\nIncentive", "in",   "input",   9,  MONEY),
    ("Other\nAllow",        "in",   "input",   8,  MONEY),
    ("Work\nDays",          "in",   "input",   6,  "0"),
    ("Unpaid\nDays",        "in",   "input",   6,  "0.0"),
    ("OT\nHours",           "in",   "input",   6,  "0.00"),
    ("OT\nRate",            "in",   "input",   6,  MONEY),
    ("Other\nDeduct",       "in",   "input",   8,  MONEY),
    ("Extra PCB\nRelief/yr","in",   "input",   9,  MONEY0),
    ("PCB\nOverride",       "in",   "input",   8,  MONEY),
    ("Prorated\nBasic",     "calc", "formula", 9,  MONEY),
    ("Gross\n(EPF wage)",   "calc", "formula", 10, MONEY),
    ("OT Pay",              "calc", "formula", 9,  MONEY),
    ("Total\nRemun.",       "calc", "formula", 10, MONEY),
    ("EPF\nEmployee",       "calc", "formula", 9,  MONEY),
    ("EPF\nEmployer",       "calc", "formula", 9,  MONEY),
    ("SOCSO\nEmployee",     "calc", "formula", 9,  MONEY),
    ("SOCSO\nEmployer",     "calc", "formula", 9,  MONEY),
    ("EIS\nEmployee",       "calc", "formula", 8,  MONEY),
    ("EIS\nEmployer",       "calc", "formula", 8,  MONEY),
    ("Charge.\nIncome/yr",  "calc", "formula", 10, MONEY0),
    ("PCB\nEstimate",       "calc", "formula", 9,  MONEY),
    ("PCB\nApplied",        "calc", "formula", 9,  MONEY),
    ("Total EE\nDeduct",    "calc", "formula", 9,  MONEY),
    ("NET SALARY\n(take-home)", "calc", "formula", 12, MONEY),
    ("Employer\nStatutory", "calc", "formula", 10, MONEY),
    ("EMPLOYER\nTOTAL COST","calc", "formula", 12, MONEY),
]
# Map header -> column letter for readable formula building
L = {c[0]: get_column_letter(i + 1) for i, c in enumerate(COLS)}


def cl(name):  # column letter by header
    return L[name]


TITLE_ROW = 1
MONTH_ROW = 2
LEGEND_ROW = 3
HDR_ROW = 4
DATA_START = 5

ncols = len(COLS)
last_col = get_column_letter(ncols)

# Title & month
otpay_col = cl("OT Pay")
totrem_col = cl("Total\nRemun.")
pay.merge_cells(f"A{TITLE_ROW}:{otpay_col}{TITLE_ROW}")
t = pay.cell(TITLE_ROW, 1, "MONTHLY PAYROLL — MALAYSIA")
t.font = TITLE
pay.merge_cells(f"{totrem_col}{TITLE_ROW}:{last_col}{TITLE_ROW}")

pay.cell(MONTH_ROW, 1, "Pay month:")
pay.cell(MONTH_ROW, 1).font = Font(name=FONT, size=10, bold=True)
mc = pay.cell(MONTH_ROW, 2, "June 2026")
mc.font = BLUE
mc.fill = FILL_INPUT
pay.cell(MONTH_ROW, 4, "Statutory contributions auto-calculated from the 'Rates' sheet.").font = NOTE

# Legend
legend = ("LEGEND:   blue = you type it (inputs)      black = auto-formula      "
          "green = pulled from Rates sheet      yellow = VERIFY against image")
pay.cell(LEGEND_ROW, 1, legend).font = Font(name=FONT, size=9, italic=True, color="C00000")

# Header row
for i, (h, group, kind, width, fmt) in enumerate(COLS, start=1):
    c = pay.cell(HDR_ROW, i, h)
    c.font = HDR
    c.alignment = CTR
    c.border = BORDER
    c.fill = {"emp": FILL_HDR, "in": FILL_HDR_IN, "calc": FILL_HDR_CALC}[group]
    pay.column_dimensions[get_column_letter(i)].width = width

# ---------------------------------------------------------------------------
# Employee / June-2026 data.
#   Fields: id, name, bank, acct, type, age60, epf, socso,
#           basic, petrol, incentive, dishwash, other, workdays, unpaid,
#           ot_hours, ot_rate, other_deduct, extra_relief, pcb_override
#   verify : set of column-headers to flag yellow + note text
# ---------------------------------------------------------------------------
def emp(id, name, bank, acct, type, epf, socso, basic=0, petrol=0, incentive=0,
        dishwash=0, other=0, workdays=26, unpaid=0, ot_hours=0, ot_rate=0,
        other_deduct=0, extra_relief=0, pcb_override=None, age60="N",
        verify=None, note=None):
    return dict(id=id, name=name, bank=bank, acct=acct, type=type, age60=age60,
                epf=epf, socso=socso, basic=basic, petrol=petrol,
                incentive=incentive, dishwash=dishwash, other=other,
                workdays=workdays, unpaid=unpaid, ot_hours=ot_hours,
                ot_rate=ot_rate, other_deduct=other_deduct,
                extra_relief=extra_relief, pcb_override=pcb_override,
                verify=verify or {}, note=note)


V = "verify"   # convenience
employees = [
    # --- Permanent / salaried (printed rows in the image) ------------------
    emp("ON00010", "cheong fong cheng", "Cimb", "7000986177", "Permanent", "Y", "Y",
        basic=5000,
        verify={"Petrol\nAllow": "Image shows a petrol tick (no amount) — enter RM if applicable."}),
    emp("ON00011", "mohd shah putra danial", "Maybank", "112205095227", "Permanent", "Y", "Y",
        basic=3200,
        verify={"Basic": "Basic cell was whited-out in the June image — verify amount / whether paid."}),
    emp("ON00012", "mohd hafizal bin azahar", "Public Bank", "4798247316", "Permanent", "Y", "Y",
        basic=3000, incentive=200, ot_rate=15,
        verify={"OT\nHours": "Image note '+200 for OT' — enter OT hours (rate 15) or use PCB/OT as needed."}),
    emp("ON00013", "Nur Azaharina Binti Azahar", "Bank Muamalat", "3060002590765", "Permanent", "Y", "Y",
        basic=2200,
        verify={"Basic": "Basic cell was whited-out in the June image — verify amount / whether paid."}),
    emp("ON00015", "Chin Ching Hua", "Public Bank", "5016571402", "Permanent", "Y", "Y",
        basic=2200, ot_hours=25.5, ot_rate=15,
        verify={"Basic": "Basic whited-out in image — verify.",
                "OT\nHours": "Image shows OT 25.5 x 15 = 382.50 near this row — confirm it is this employee."}),
    emp("ON00022", "Mohd Aminludin Bin Yahya", "Maybank", "164061840677", "Permanent", "Y", "Y",
        basic=2400, incentive=200, ot_hours=22.5, ot_rate=15,
        verify={"OT\nHours": "Image shows OT 22.5 x 15 = 337.50 near this row — confirm it is this employee."}),
    emp("ON00023", "Nur Alya Adriana", "Maybank", "564593318369", "Permanent", "Y", "Y",
        basic=2100,
        verify={"Basic": "Basic cell was whited-out in the June image — verify amount / whether paid."}),
    emp("ON00028", "Tan Kui Thean", "Maybank", "164061854800", "Permanent", "Y", "Y",
        basic=3200, ot_rate=15,
        verify={"Petrol\nAllow": "Image shows a petrol tick — enter RM if applicable."}),
    # --- Part-time / hourly (printed, lower rows) --------------------------
    emp("ON00030", "HOU CHING", "Public Bank", "5102691912", "Part-time", "N", "N",
        ot_rate=8,
        verify={"OT\nHours": "Part-timer paid hours x RM8 — enter June hours from image."}),
    emp("ON00031", "Aron cheng Eng", "Cimb", "7657001861", "Part-time", "N", "N",
        ot_rate=8,
        verify={"OT\nHours": "Part-timer paid hours x RM8 — enter June hours from image."}),
    # --- Hand-written part-time / hourly block (clearly legible) -----------
    emp("", "Liew Chen Hao", "Cimb", "", "Part-time", "N", "N", ot_hours=51, ot_rate=8),
    emp("", "Tai Jun Xi", "Cash", "", "Part-time", "N", "N", ot_hours=41, ot_rate=8),
    emp("ON00055", "Nay Lin Zaw", "Merchantrade", "500001618970", "Part-time", "N", "N",
        basic=2300, workdays=26,
        note="Foreign worker (per May sheet). SOCSO Employment-Injury applies to foreign "
             "workers; EPF 2% mandatory from Oct 2025 — toggle 'Apply' columns to Y if needed.",
        verify={"Basic": "Image shows 2300 — confirm monthly wage / any unpaid days."}),
    emp("", "Ong Peir Ann", "Public Bank", "5097798113", "Part-time", "N", "N", ot_hours=210, ot_rate=8),
    emp("", "Thulasi Raj", "Cimb", "7651498174", "Part-time", "N", "N", ot_hours=4, ot_rate=8),
    emp("", "Shannen Marie Gomez", "Maybank", "", "Part-time", "N", "N", ot_hours=7, ot_rate=8),
    emp("", "Soh Jia Mon", "Public Bank", "5089470408", "Part-time", "N", "N", ot_hours=28, ot_rate=8),
    emp("", "Chee Wei Hong", "Maybank", "", "Part-time", "N", "N", ot_hours=87, ot_rate=8),
]

# Header -> field name for input columns
INPUT_FIELD = {
    "Type": "type", "Age\n>= 60?": "age60", "Apply\nEPF?": "epf",
    "Apply\nSOCSO\n+EIS?": "socso", "Basic": "basic", "Petrol\nAllow": "petrol",
    "Incentive": "incentive", "Dishwash\nIncentive": "dishwash",
    "Other\nAllow": "other", "Work\nDays": "workdays", "Unpaid\nDays": "unpaid",
    "OT\nHours": "ot_hours", "OT\nRate": "ot_rate", "Other\nDeduct": "other_deduct",
    "Extra PCB\nRelief/yr": "extra_relief", "PCB\nOverride": "pcb_override",
}
TEXT_FIELD = {"EMP_ID": "id", "Name": "name", "Bank Name": "bank", "Bank Acct No": "acct"}


def formula(colname, row, e):
    """Return the Excel formula string for a calc column at the given row."""
    B, Pt, In, Dw, Ot = cl("Basic"), cl("Petrol\nAllow"), cl("Incentive"), cl("Dishwash\nIncentive"), cl("Other\nAllow")
    WD, UD = cl("Work\nDays"), cl("Unpaid\nDays")
    OH, OR_ = cl("OT\nHours"), cl("OT\nRate")
    ProB, Gross, OTPay, TotRem = cl("Prorated\nBasic"), cl("Gross\n(EPF wage)"), cl("OT Pay"), cl("Total\nRemun.")
    epfE, epfR = cl("EPF\nEmployee"), cl("EPF\nEmployer")
    socE, socR = cl("SOCSO\nEmployee"), cl("SOCSO\nEmployer")
    eisE, eisR = cl("EIS\nEmployee"), cl("EIS\nEmployer")
    ci = cl("Charge.\nIncome/yr")
    pcbEst, pcbApp = cl("PCB\nEstimate"), cl("PCB\nApplied")
    totDed, net = cl("Total EE\nDeduct"), cl("NET SALARY\n(take-home)")
    erStat = cl("Employer\nStatutory")
    aEPF, aSOC = cl("Apply\nEPF?"), cl("Apply\nSOCSO\n+EIS?")
    age = cl("Age\n>= 60?")
    xrel, pcbov, odd = cl("Extra PCB\nRelief/yr"), cl("PCB\nOverride"), cl("Other\nDeduct")
    A = addr  # rates addresses

    g_wage = f"MIN({Gross}{row},{A['ceiling']})"   # wage used for SOCSO/EIS lookup
    f = {
        "Prorated\nBasic":
            f"=IF({WD}{row}=0,{B}{row},{B}{row}-{UD}{row}*({B}{row}/{WD}{row}))",
        "Gross\n(EPF wage)":
            f"={ProB}{row}+{Pt}{row}+{In}{row}+{Dw}{row}+{Ot}{row}",
        "OT Pay":
            f"={OH}{row}*{OR_}{row}",
        "Total\nRemun.":
            f"={Gross}{row}+{OTPay}{row}",
        "EPF\nEmployee":
            f'=IF({aEPF}{row}="Y",ROUNDUP({Gross}{row}*{A["epf_emp"]},0),0)',
        "EPF\nEmployer":
            f'=IF({aEPF}{row}="Y",ROUNDUP({Gross}{row}*IF({TotRem}{row}>{A["epf_thresh"]},{A["epf_er_hi"]},{A["epf_er_lo"]}),0),0)',
        "SOCSO\nEmployee":
            f'=IF({aSOC}{row}="Y",IF({age}{row}="Y",0,VLOOKUP({g_wage},{A["soc_tbl"]},{SOC_C1_EMP},TRUE)),0)',
        "SOCSO\nEmployer":
            f'=IF({aSOC}{row}="Y",IF({age}{row}="Y",VLOOKUP({g_wage},{A["soc_tbl"]},{SOC_C2_ER},TRUE),VLOOKUP({g_wage},{A["soc_tbl"]},{SOC_C1_ER},TRUE)),0)',
        "EIS\nEmployee":
            f'=IF(AND({aSOC}{row}="Y",{age}{row}<>"Y"),VLOOKUP({g_wage},{A["soc_tbl"]},{EIS_COL},TRUE),0)',
        "EIS\nEmployer":
            f'=IF(AND({aSOC}{row}="Y",{age}{row}<>"Y"),VLOOKUP({g_wage},{A["soc_tbl"]},{EIS_COL},TRUE),0)',
        "Charge.\nIncome/yr":
            # annual chargeable income = (EPF wage x12) - personal - EPF relief - extra reliefs
            (f"=MAX(0,{Gross}{row}*12-{A['relief_personal']}"
             f"-MIN({epfE}{row}*12,{A['relief_epf']})-{xrel}{row})"),
        "PCB\nEstimate":
            # progressive tax on chargeable income, less rebate, spread over 12 months
            (f"=MAX(0,MROUND((VLOOKUP({ci}{row},{A['tax_tbl']},3,TRUE)"
             f"+({ci}{row}-VLOOKUP({ci}{row},{A['tax_tbl']},1,TRUE))"
             f"*VLOOKUP({ci}{row},{A['tax_tbl']},2,TRUE)"
             f"-IF({ci}{row}<=35000,{A['rebate']},0))/12,0.05))"),
        "PCB\nApplied":
            f'=IF({pcbov}{row}="",{pcbEst}{row},{pcbov}{row})',
        "Total EE\nDeduct":
            f"={epfE}{row}+{socE}{row}+{eisE}{row}+{pcbApp}{row}+{odd}{row}",
        "NET SALARY\n(take-home)":
            f"={TotRem}{row}-{totDed}{row}",
        "Employer\nStatutory":
            f"={epfR}{row}+{socR}{row}+{eisR}{row}",
        "EMPLOYER\nTOTAL COST":
            f"={TotRem}{row}+{erStat}{row}",
    }
    return f[colname]


# Write data rows
row = DATA_START
for e in employees:
    for i, (h, group, kind, width, fmt) in enumerate(COLS, start=1):
        c = pay.cell(row, i)
        c.border = BORDER
        c.font = BLACK
        if fmt:
            c.number_format = fmt
        if h in TEXT_FIELD:
            c.value = e[TEXT_FIELD[h]]
            c.alignment = LEFT if h in ("Name", "Bank Name") else CTR
        elif h in INPUT_FIELD:
            val = e[INPUT_FIELD[h]]
            if h == "PCB\nOverride":
                c.value = "" if val is None else val
            else:
                c.value = val
            c.font = BLUE
            c.fill = FILL_INPUT
            c.alignment = CTR if h in ("Type", "Age\n>= 60?", "Apply\nEPF?",
                                       "Apply\nSOCSO\n+EIS?", "Work\nDays",
                                       "Unpaid\nDays") else RIGHT
            # verify flag
            if h in e["verify"]:
                c.fill = FILL_VERIFY
                c.comment = Comment(e["verify"][h], "Payroll build")
        elif group == "calc":
            c.value = formula(h, row, e)
            c.alignment = RIGHT
            # green tint for columns that pull statutory rates via lookup
            if h in ("EPF\nEmployee", "EPF\nEmployer", "SOCSO\nEmployee",
                     "SOCSO\nEmployer", "EIS\nEmployee", "EIS\nEmployer",
                     "PCB\nEstimate"):
                c.font = GREEN
    # employee-level note on the Name cell
    if e["note"]:
        pay.cell(row, 2).comment = Comment(e["note"], "Payroll build")
    # emphasise NET SALARY column
    pay.cell(row, COLS.index(("NET SALARY\n(take-home)", "calc", "formula", 12, MONEY)) + 1).font = \
        Font(name=FONT, size=10, bold=True, color="006100")
    row += 1

DATA_END = row - 1

# ---------------------------------------------------------------------------
# Totals row
# ---------------------------------------------------------------------------
trow = row + 0
pay.cell(trow, 1, "TOTALS").font = Font(name=FONT, size=10, bold=True)
for i, (h, group, kind, width, fmt) in enumerate(COLS, start=1):
    c = pay.cell(trow, i)
    c.fill = FILL_TOTAL
    c.border = BORDER
    if group == "calc" or h in ("Basic", "Petrol\nAllow", "Incentive",
                                "Dishwash\nIncentive", "Other\nAllow", "Other\nDeduct"):
        letter = get_column_letter(i)
        c.value = f"=SUM({letter}{DATA_START}:{letter}{DATA_END})"
        c.font = Font(name=FONT, size=10, bold=True, color="000000")
        c.number_format = MONEY
        c.alignment = RIGHT
pay.cell(trow, 1).fill = FILL_TOTAL

# ---------------------------------------------------------------------------
# Data validations (dropdowns) for Y/N and Type
# ---------------------------------------------------------------------------
def add_dv(colheader, formula1):
    letter = cl(colheader)
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(f"{letter}{DATA_START}:{letter}{DATA_END}")
    pay.add_data_validation(dv)

add_dv("Type", '"Permanent,Part-time,Foreign"')
add_dv("Age\n>= 60?", '"Y,N"')
add_dv("Apply\nEPF?", '"Y,N"')
add_dv("Apply\nSOCSO\n+EIS?", '"Y,N"')

# Freeze panes: keep header + first two columns visible
pay.freeze_panes = f"C{DATA_START}"
pay.sheet_view.showGridLines = False
pay.row_dimensions[HDR_ROW].height = 42

# ===========================================================================
# 3. INSTRUCTIONS SHEET
# ===========================================================================
ins = wb.create_sheet("How to use")
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 110
lines = [
    ("MONTHLY PAYROLL WORKBOOK — HOW TO USE", "title"),
    ("", ""),
    ("This workbook turns your monthly payroll image into a calculated payslip run. "
     "You only type the blue cells; EPF (KWSP), SOCSO, EIS and PCB are derived automatically.", "n"),
    ("", ""),
    ("EACH MONTH — 4 STEPS", "h"),
    ("1.  Update 'Pay month' (Payroll!B2) and enter each employee's inputs from the image: "
     "Basic, Petrol/Incentive/Dishwash/Other allowances, Work Days, Unpaid Days, OT Hours and OT Rate.", "n"),
    ("2.  Set the three switches per person:  Type (Permanent/Part-time/Foreign), "
     "'Apply EPF?' and 'Apply SOCSO+EIS?' (Y/N), and 'Age >= 60?' for the SOCSO category.", "n"),
    ("3.  Yellow cells are best-effort reads from this month's image — open the image and confirm each one.", "n"),
    ("4.  Read the results:  NET SALARY (take-home) per employee, and EMPLOYER TOTAL COST "
     "(gross + employer statutory). The TOTALS row foots every money column.", "n"),
    ("", ""),
    ("HOW EACH CONTRIBUTION IS DERIVED", "h"),
    ("EPF / KWSP  — Employee = 11% of EPF wage (rounded up to the next RM). "
     "Employer = 13% if total remuneration <= RM5,000, else 12%. Rates live on the 'Rates' sheet.", "n"),
    ("SOCSO / PERKESO — read from the official RM100-band table (ceiling RM6,000). "
     "Category 1 (age < 60): employee 0.5% + employer 1.75% of the band midpoint. "
     "Category 2 (age >= 60): employer 1.25% only, employee 0.", "n"),
    ("EIS — 0.2% each side from the same band table (ceiling RM6,000); not applied at age >= 60.", "n"),
    ("PCB (income tax, MTD) — an ESTIMATE only: annualised chargeable income "
     "(EPF wage x 12 less personal relief RM9,000, EPF relief up to RM4,000, and any 'Extra PCB Relief/yr' "
     "you enter for spouse/children) taxed on YA2024/25 resident brackets, less the RM400 rebate, divided by 12. "
     "For the exact figure use LHDN e-PCB / e-Jadual PCB and type it into the 'PCB Override' column — the override always wins.", "n"),
    ("", ""),
    ("WAGE DEFINITIONS (match the previous May workbook)", "h"),
    ("Gross (EPF wage) = prorated Basic + Petrol + Incentive + Dishwash + Other allowances. Overtime is NOT part of the EPF wage.", "n"),
    ("Prorated Basic = Basic - Unpaid Days x (Basic / Work Days).  Default Work Days = 26.", "n"),
    ("SOCSO/EIS use the same Gross wage, capped at the RM6,000 ceiling.", "n"),
    ("Note: petrol/travel allowance is technically EPF/SOCSO-exempt; it is included here to match your existing sheet. "
     "Move it to a non-statutory column if you want it excluded.", "n"),
    ("", ""),
    ("PART-TIME / HOURLY STAFF", "h"),
    ("For hourly staff, leave Basic = 0, put hours in 'OT Hours' and the hourly rate in 'OT Rate'; "
     "pay = hours x rate. 'Apply EPF?' and 'Apply SOCSO+EIS?' are defaulted to N — set to Y if they are covered.", "n"),
    ("Foreign workers (e.g. Nay Lin Zaw): SOCSO Employment-Injury applies, and EPF becomes mandatory at 2% from Oct 2025. "
     "Switch the 'Apply' columns to Y and adjust rates on the 'Rates' sheet as needed.", "n"),
    ("", ""),
    ("COLOUR KEY", "h"),
    ("Blue = you type it     Black = auto-formula     Green = pulled from Rates sheet     Yellow = verify against the image", "n"),
    ("", ""),
    ("SOURCES", "h"),
    ("PERKESO rate of contribution (SOCSO/EIS): https://www.perkeso.gov.my/en/rate-of-contribution.html", "n"),
    ("KWSP/EPF Third Schedule: https://www.kwsp.gov.my/en/employer/responsibilities/mandatory-contribution", "n"),
    ("LHDN individual tax rates (YA2024/2025): https://www.hasil.gov.my/en/individual/", "n"),
]
rr = 1
for text, kind in lines:
    c = ins.cell(rr, 2, text)
    if kind == "title":
        c.font = TITLE
    elif kind == "h":
        c.font = Font(name=FONT, size=11, bold=True, color="1F3864")
    else:
        c.font = Font(name=FONT, size=10, color="000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")
    rr += 1

# ===========================================================================
wb.calculation.fullCalcOnLoad = True    # force Excel/LibreOffice to recompute
wb.save(OUT)
print(f"Wrote {OUT}: sheets={wb.sheetnames}, employees={len(employees)}, "
      f"data rows {DATA_START}-{DATA_END}, SOCSO table {addr['soc_tbl']}")
