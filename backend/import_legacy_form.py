"""Phase 3 — import the legacy Google Form export into the employee master.

The 58 registration rows already exist as ``employees`` (seeded by seed.py,
matched here by emp-code order / name+email). This script ENRICHES them:

* raw row -> application_submissions (verbatim JSON, source='google_form_import')
* employees: address, identity_type/NRIC recovery, phone normalization,
  status ('active' for the current roster, else 'pending_review') + history
* bank_accounts (verified=false), emergency_contacts, employee_documents
  (Drive URLs in source_url only)
* hr_review_flags for every recovery/guess — nothing is fixed silently

Idempotent: keyed on application_submissions.reference_no = 'GF-<row>'.
Re-runs update in place and never duplicate.

Usage:
    DATABASE_URL=postgresql+psycopg://... python import_legacy_form.py [xlsx]
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import Employee

DEFAULT_FORM = Path(__file__).resolve().parent / "employee-registration-form.xlsx"

# Staff active for the current period (mirrors seed.py's roster).
ACTIVE_ROSTER = [
    "cheong fong cheng", "tan kui thean", "mohd aminludin bin yahya",
    "mohd hafizal bin azahar", "liew chen hao", "tai jun xi", "nay lin zaw",
    "ong peir ann", "thulasi raj murugan", "shannen gomes", "soh jia man",
    "chee wei hong",
]

HEADERS = [
    "Timestamp", "Name", "Email", "Residential Address", "Phone number",
    "NRIC", "Upload NRIC Front", "Upload NRIC Back", "DOB", "Bank Name",
    "Bank Account", "Job Position Applied For", "Emergency Contact Name",
    "Working Experience", "Education Background", "Expected Salary",
    "Comments", "Employee ID", "Upload Khusus Makanan Certification",
    "Upload Typhoid Vaccination Proof", "Typhoid Vaccination Expiry Date",
    "Education Transcript", "Referral Source", "Skills",
    "Emergency Contact Number", "Emergency Contact Relationship",
    "Profile Picture",
]

DOC_COLS = {  # column index (1-based) -> doc_type
    7: "nric_front", 8: "nric_back", 19: "food_handler_cert",
    20: "typhoid_proof", 22: "education_transcript", 27: "profile_photo",
}

POSITIONS = ["waiter", "waitress", "chef", "kitchen", "supervisor",
             "outlet manager", "manager", "barista", "cashier"]


def _norm(s: str) -> str:
    return " ".join((s or "").lower().split())


def _cellstr(v) -> str:
    """Render a cell as text, expanding float/scientific notation losslessly."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return repr(v)
    return str(v).strip()


def _json_cell(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()
    return v


# --- recovery helpers --------------------------------------------------------

def _valid_nric(digits: str) -> bool:
    if not re.fullmatch(r"\d{12}", digits):
        return False
    yy, mm, dd = int(digits[0:2]), int(digits[2:4]), int(digits[4:6])
    try:
        dt.date(2000 + yy if yy < 30 else 1900 + yy, mm, dd)
        return True
    except ValueError:
        return False


def recover_identity(raw, was_numeric: bool):
    """-> (identity_type, identity_no|None, flags[(type,severity,note)])."""
    s = _cellstr(raw)
    if not s or s in ("-",):
        return "nric", None, [("invalid_identity_no", "blocker", "identity number missing")]
    if s.upper() == "UNHCR":
        return "unhcr", None, [("work_authorization_review", "blocker",
                                "UNHCR holder — verify UNHCR card + work authorization")]
    dashless = re.sub(r"[\s-]", "", s)
    if re.fullmatch(r"[A-Za-z]{1,2}\d{6,9}", dashless):
        return "passport", dashless.upper(), [("work_authorization_review", "blocker",
                                               "passport holder — verify work permit (PLKS) before activation")]
    if dashless.isdigit():
        flags = []
        if len(dashless) == 11 and _valid_nric("0" + dashless):
            dashless = "0" + dashless
            flags.append(("invalid_identity_no", "info",
                          "leading zero restored from numeric cell — verify against NRIC document"))
        if _valid_nric(dashless):
            if was_numeric:
                flags.append(("invalid_identity_no", "info",
                              "recovered from numeric/scientific-notation cell — verify"))
            return "nric", dashless, flags
        return "nric", dashless, [("invalid_identity_no", "blocker",
                                   f"{len(dashless)}-digit value fails NRIC YYMMDD check")]
    return "nric", None, [("invalid_identity_no", "blocker",
                           "unparseable identity value — see submission payload")]


def normalize_phone(raw):
    """-> (normalized|None, flags). TEXT in, TEXT out."""
    s = _cellstr(raw)
    if not s:
        return None, []
    digits = re.sub(r"\D", "", s)
    flags = []
    if isinstance(raw, (int, float)):
        # floats lost the leading 0 (and sometimes arrived as 60xxxxxxxxx)
        if digits.startswith("1"):
            digits = "60" + digits      # 0 lost from 01x-xxxxxxx
        flags.append(("corrupted_phone", "warning",
                      "phone arrived as a numeric cell — leading digits may be lost; verify"))
    if digits.startswith("600"):
        digits = "60" + digits[3:]
    elif digits.startswith("0"):
        digits = "60" + digits[1:]
    elif not digits.startswith("60"):
        digits = "60" + digits
    if not 10 <= len(digits) <= 13:
        flags.append(("corrupted_phone", "warning", f"unusual phone length after normalization: +{digits}"))
    return "+" + digits, flags


def parse_bank(bank_raw, account_raw, name: str):
    """Handle swapped/embedded bank name & account. -> (bank, account, flags)."""
    bank, account = _cellstr(bank_raw), _cellstr(account_raw)
    flags = []
    if bank in ("-", "") and account in ("-", ""):
        return None, None, [("corrupted_bank_account", "blocker", "no bank details provided")]
    # account text sitting in the bank column (e.g. bank='4576935026', account='PUBLIC')
    if re.fullmatch(r"[\d .-]{7,}", bank) and not re.search(r"\d{6,}", account):
        bank, account = account, bank
        flags.append(("corrupted_bank_account", "warning", "bank name and account number appear swapped — verify"))
    # bank column holds the account holder's own name (e.g. 'Tan Wei Yuan')
    if _norm(bank) == _norm(name):
        flags.append(("corrupted_bank_account", "warning",
                      "bank-name column contains the applicant's name, not a bank — verify"))
    # 'PUBLIC BANK 4922328724' style: bank embedded in the account cell
    m = re.match(r"([A-Za-z][A-Za-z ]{3,}?)\s*(\d[\d ]{6,})$", account)
    if m and not re.search(r"\d", bank):
        pass  # bank cell already fine
    elif m:
        account = m.group(2)
        flags.append(("corrupted_bank_account", "warning", "bank name embedded in account cell — split; verify"))
    digits = re.sub(r"\D", "", account)
    if isinstance(account_raw, float):
        flags.append(("corrupted_bank_account", "blocker",
                      "account number came from a numeric cell — digits may be corrupted; re-confirm with employee"))
    if len(digits) >= 15:
        flags.append(("corrupted_bank_account", "blocker",
                      f"{len(digits)}-digit account at float precision limit — almost certainly corrupted"))
    elif digits and not 6 <= len(digits) <= 16:
        flags.append(("corrupted_bank_account", "blocker", f"account length {len(digits)} fails basic check"))
    return (bank or None), (re.sub(r"\s", "", account) or None), flags


def parse_position(raw):
    s = _cellstr(raw)
    ltxt = s.lower()
    etype = "part_time" if "part time" in ltxt or "part-time" in ltxt else None
    first = re.split(r"[,/]| and ", s)[0].strip()
    if _norm(first) in ("part time", "part-time"):
        first = None
    return first or None, etype, s


def parse_dob(raw, submitted: dt.datetime):
    if raw is None:
        return None, [("suspicious_dob", "warning", "DOB missing")]
    if isinstance(raw, str):
        return None, [("suspicious_dob", "warning", f"unparseable DOB text: {raw!r}")]
    d = raw.date() if isinstance(raw, dt.datetime) else raw
    if d == submitted.date() or abs((submitted.date() - d).days) < 30:
        return None, [("suspicious_dob", "warning", "DOB equals/near submission date — likely form error")]
    age = (submitted.date() - d).days / 365.25
    if age < 15 or age > 70:
        return None, [("suspicious_dob", "warning", f"implausible age {age:.0f} at submission")]
    return d, []


def salary_hint(raw):
    s = _cellstr(raw).lower()
    if not s or s in ("-",):
        return None, [("ambiguous_salary", "warning", "no expected salary given")]
    hourly = any(h in s for h in ("hour", "/hr", "per hr", "jam", "ringgit"))
    monthly = bool(re.search(r"\b[12]\d{3}\b|\dk", s)) or "basic" in s or "month" in s
    flags = []
    if hourly and monthly:
        flags.append(("ambiguous_salary", "warning", f"mixes hourly and monthly: {s!r}"))
    if not hourly and not monthly:
        flags.append(("ambiguous_salary", "warning", f"unparseable expectation: {s!r}"))
    return ("hourly" if hourly else "monthly" if monthly else None), flags


def split_contact(name_raw, number_raw, rel_raw):
    """Early rows mixed phone into the contact-name field."""
    name, number, rel = _cellstr(name_raw), _cellstr(number_raw), _cellstr(rel_raw)
    if not number:
        m = re.search(r"(\+?[\d][\d\s()-]{7,})", name)
        if m:
            number = m.group(1).strip()
            name = name.replace(m.group(1), "").strip(" ,()-") or None
    if not rel:
        m = re.search(r"\(([^)]+)\)", name or "")
        if m and not re.search(r"\d", m.group(1)):
            rel = m.group(1)
    return name or None, number or None, rel or None


# --- main --------------------------------------------------------------------

def import_worksheet(db, ws) -> dict:
    """Import/refresh every form row. Reusable by the CLI and the HR
    sync-google-sheet endpoint. Employees that HR has already moved past
    review (active/resigned/rejected/…) are left untouched apart from the
    raw submission upsert."""
    stats = {"rows": 0, "matched": 0, "created": 0, "skipped_settled": 0}
    flag_counts: dict[str, int] = {}
    urgent: list[str] = []

    if True:
        db.execute(text("set onni.bypass_lifecycle = '1'"))
        employees = db.scalars(select(Employee)).all()
        by_key = {}
        for e in employees:
            by_key[_norm(e.name)] = e
            if e.email:
                by_key.setdefault(e.email.lower().strip(), e)

        seen_identity: dict[str, int] = {}
        seen_phone: dict[str, int] = {}
        seen_email: dict[str, int] = {}

        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(r, c).value for c in range(1, 28)]
            if not any(cells):
                continue
            stats["rows"] += 1
            ref = f"GF-{r:03d}"
            payload = {HEADERS[i]: _json_cell(cells[i]) for i in range(27) if cells[i] is not None}
            ts = cells[0] if isinstance(cells[0], dt.datetime) else dt.datetime.now()
            name = _cellstr(cells[1])
            email = _cellstr(cells[2]).lower() or None
            flags: list[tuple[str, str, str]] = []

            if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.(?!con$)[a-z]{2,}", email):
                flags.append(("invalid_email", "warning", f"email looks wrong: {email}"))

            address = _cellstr(cells[3]) or None
            if address and address.isdigit():
                flags.append(("data_issue", "warning", f"residential address is just a number: {address!r}"))
                address = None

            phone, f2 = normalize_phone(cells[4])
            flags += f2
            idtype, idno, f3 = recover_identity(cells[5], isinstance(cells[5], (int, float)))
            flags += f3
            dob, f4 = parse_dob(cells[8], ts)
            flags += f4
            bank, account, f5 = parse_bank(cells[9], cells[10], name)
            flags += f5
            position, etype_hint, pos_raw = parse_position(cells[11])
            pay_hint, f6 = salary_hint(cells[15])
            flags += f6
            payload["_derived"] = {"pay_type_hint": pay_hint, "position_raw": pos_raw}

            # --- find or create the employee ---------------------------------
            emp = by_key.get(_norm(name)) or (email and by_key.get(email))
            if emp and emp.status not in ("applicant", "pending_review"):
                # HR already settled this person (active/resigned/rejected...):
                # keep the raw submission current, change nothing else.
                stats["skipped_settled"] += 1
                db.execute(text("""
                    insert into application_submissions (source, reference_no, payload, employee_id, processed_at, created_at)
                    values ('google_form_import', :ref, :payload, :eid, now(), :ts)
                    on conflict (reference_no) do update set payload = excluded.payload
                """), {"ref": ref, "payload": json.dumps(payload, ensure_ascii=False, default=str),
                       "eid": emp.id, "ts": ts})
                continue
            if emp:
                stats["matched"] += 1
            else:
                code = db.scalar(text(
                    "select 'ONNI' || lpad((coalesce(max(substring(emp_code from '\\d+')::int),0)+1)::text, 4, '0') from employees"))
                emp = Employee(emp_code=code, name=name)
                db.add(emp)
                db.flush()
                stats["created"] += 1

            emp.email = emp.email or email
            emp.phone = phone or emp.phone
            emp.nric = idno or emp.nric
            emp.identity_type = idtype
            emp.residential_address = address
            emp.nationality = "MY" if idtype == "nric" else "non-MY"
            emp.dob = dob or emp.dob
            emp.position = emp.position or position
            emp.is_foreign = idtype in ("passport", "unhcr")

            is_active = _norm(name) in (_norm(n) for n in ACTIVE_ROSTER)
            target_status = "active" if is_active else "pending_review"

            # duplicates across the batch
            for val, seen, label in ((idno, seen_identity, "identity_no"),
                                     (phone, seen_phone, "phone"),
                                     (email, seen_email, "email")):
                if val:
                    if val in seen and seen[val] != emp.id:
                        flags.append(("duplicate_suspect", "blocker",
                                      f"same {label} as employee id {seen[val]}"))
                    seen.setdefault(val, emp.id)

            db.flush()
            eid = emp.id

            # --- submission (idempotency anchor) ------------------------------
            db.execute(text("""
                insert into application_submissions (source, reference_no, payload, employee_id, processed_at, created_at)
                values ('google_form_import', :ref, :payload, :eid, now(), :ts)
                on conflict (reference_no) do update
                  set payload = excluded.payload, employee_id = excluded.employee_id
            """), {"ref": ref, "payload": json.dumps(payload, ensure_ascii=False,
                                                     default=str), "eid": eid, "ts": ts})

            # wipe our previously imported satellite rows for clean re-runs
            for tbl in ("bank_accounts", "emergency_contacts", "employee_documents"):
                db.execute(text(f"delete from {tbl} where employee_id = :eid"), {"eid": eid})
            db.execute(text("delete from hr_review_flags where employee_id = :eid and raised_by = 'import'"),
                       {"eid": eid})

            # --- status + history ---------------------------------------------
            db.execute(text("delete from employee_status_history where employee_id = :eid"), {"eid": eid})
            db.execute(text("""
                insert into employee_status_history (employee_id, status, effective_date, reason, changed_by)
                values (:eid, :st, :d, 'legacy google-form import', 'import')
            """), {"eid": eid, "st": target_status, "d": ts.date()})
            if is_active and not emp.hire_date:
                emp.hire_date = ts.date()

            # --- satellites ----------------------------------------------------
            if bank or account:
                db.execute(text("""
                    insert into bank_accounts (employee_id, bank_name, account_no, account_holder_name, verified)
                    values (:eid, :b, :a, :h, false)
                """), {"eid": eid, "b": bank or "UNKNOWN", "a": account or "", "h": name})
                flags.append(("unverified_bank_account", "blocker",
                              "imported bank account — must be re-confirmed before payroll inclusion"))

            cname, cnum, crel = split_contact(cells[12], cells[24], cells[25])
            if cname or cnum:
                db.execute(text("""
                    insert into emergency_contacts (employee_id, name, phone, relationship)
                    values (:eid, :n, :p, :r)
                """), {"eid": eid, "n": cname or "(unknown)", "p": cnum or "", "r": crel})

            typhoid_expiry = cells[20]
            if isinstance(typhoid_expiry, str):
                typhoid_expiry = None
                flags.append(("data_issue", "warning", f"unparseable typhoid expiry: {cells[20]!r}"))
            elif isinstance(typhoid_expiry, dt.datetime):
                typhoid_expiry = typhoid_expiry.date()

            any_doc = False
            for col, dtyp in DOC_COLS.items():
                url = _cellstr(cells[col - 1])
                if url.startswith("http"):
                    any_doc = True
                    db.execute(text("""
                        insert into employee_documents (employee_id, doc_type, source_url, expiry_date, verified)
                        values (:eid, :t, :u, :x, false)
                    """), {"eid": eid, "t": dtyp, "u": url,
                           "x": typhoid_expiry if dtyp == "typhoid_proof" else None})
            if any_doc:
                flags.append(("missing_document", "info",
                              "documents are Google Drive links only — file migration to storage pending"))
            for required, label in (("food_handler_cert", "food handler (Khusus Makanan) cert"),
                                    ("typhoid_proof", "typhoid vaccination proof")):
                if not _cellstr(cells[[k for k, v in DOC_COLS.items() if v == required][0] - 1]).startswith("http"):
                    flags.append(("missing_document", "warning", f"missing {label} (mandatory for F&B food handlers)"))

            if typhoid_expiry and typhoid_expiry < dt.date.today():
                flags.append(("typhoid_expired", "blocker" if is_active else "warning",
                              f"typhoid proof expired {typhoid_expiry}"))
            elif typhoid_expiry and typhoid_expiry < dt.date.today() + dt.timedelta(days=30):
                flags.append(("typhoid_expiring", "warning", f"typhoid proof expires {typhoid_expiry}"))

            for ftype, sev, note in flags:
                db.execute(text("""
                    insert into hr_review_flags (employee_id, flag_type, severity, details, raised_by)
                    values (:eid, :t, :s, :d, 'import')
                """), {"eid": eid, "t": ftype, "s": sev,
                       "d": json.dumps({"note": note, "ref": ref}, ensure_ascii=False)})
                flag_counts[ftype] = flag_counts.get(ftype, 0) + 1
                if sev == "blocker":
                    urgent.append(f"  {emp.emp_code} {name}: {ftype} — {note}")

            # commit status via direct set (bypass GUC active in this session)
            db.execute(text("update employees set status = :st, active = :act where id = :eid"),
                       {"st": target_status, "act": is_active, "eid": eid})

        db.commit()

    return {**stats, "flags": flag_counts, "urgent": urgent}


def run(xlsx: Path) -> None:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Form Responses 1"]
    with SessionLocal() as db:
        r = import_worksheet(db, ws)
    print(f"rows processed : {r['rows']}")
    print(f"matched        : {r['matched']}   created: {r['created']}   "
          f"settled (untouched): {r['skipped_settled']}")
    print("flags raised   :")
    for k, v in sorted(r["flags"].items(), key=lambda x: -x[1]):
        print(f"  {k:26s} {v}")
    print(f"urgent (blockers): {len(r['urgent'])}")
    print("\n".join(r["urgent"]))


if __name__ == "__main__":
    run(Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FORM)
