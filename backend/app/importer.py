"""Import the first batch of staff from the Google-Form registration export.

The ``Expected Salary`` column is free text (e.g. ``"RM8 per hour"``,
``"2000-2500"``, ``"3k++"``, ``"basic 2400"``), so this module makes a
best-effort guess at full-time vs part-time and the pay figure, and flags
low-confidence rows with :attr:`Employee.needs_review` for the admin to confirm.
"""

from __future__ import annotations

import datetime as dt
import re
from decimal import Decimal

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .models import Employee

# Registration-form column letters (1-based positions in the sheet).
COL = {"name": 2, "email": 3, "phone": 5, "nric": 6, "dob": 9, "bank": 10,
       "account": 11, "position": 12, "expected": 16, "reg_id": 18}

_HOURLY_HINTS = ("hour", "/hr", "per hr", "hourly", "/ hour", "phour")
_NUM = re.compile(r"(\d[\d,]*(?:\.\d+)?)\s*(k)?", re.IGNORECASE)


def _to_number(token: str, k: str) -> Decimal | None:
    try:
        val = Decimal(token.replace(",", ""))
    except Exception:
        return None
    if k:
        val *= 1000
    return val


def parse_expected_salary(raw, position: str) -> dict:
    """Infer employment type + pay figure from the free-text salary field.

    Returns dict: employment_type, basic, hourly_rate, needs_review, note.
    """
    text = ("" if raw is None else str(raw)).strip()
    pos = (position or "").lower()
    low = text.lower()
    note_bits: list[str] = []

    numbers = [n for n in (_to_number(m.group(1), m.group(2))
                           for m in _NUM.finditer(text)) if n is not None]
    hourly = any(h in low for h in _HOURLY_HINTS) or "part" in low or "part" in pos

    # Explicit "per hour" wins; also treat tiny standalone numbers as hourly.
    if not hourly and numbers and all(n <= 30 for n in numbers) and "month" not in low:
        # values like 8, 9, 10 with no monthly context are almost always hourly
        if max(numbers) <= 30 and ("rm" in low or len(text) <= 6 or "hour" in low
                                   or "/" in text):
            hourly = True

    if hourly:
        rate_candidates = [n for n in numbers if 3 <= n <= 60]
        rate = rate_candidates[0] if rate_candidates else Decimal("8")
        if not rate_candidates:
            note_bits.append("hourly rate defaulted to RM8 — confirm")
        return {"employment_type": "part_time", "basic": Decimal("0"),
                "hourly_rate": rate, "needs_review": not rate_candidates,
                "note": _joined(text, note_bits)}

    # Monthly / full-time
    monthly = [n for n in numbers if n >= 100]
    if not monthly:
        note_bits.append("could not read a monthly salary — set it manually")
        return {"employment_type": "full_time", "basic": Decimal("0"),
                "hourly_rate": Decimal("0"), "needs_review": True,
                "note": _joined(text, note_bits)}
    basic = min(monthly)   # ranges like 2000-2500 -> take the conservative lower bound
    if len(monthly) > 1:
        note_bits.append(f"range given ({'/'.join(str(int(m)) for m in monthly)}) — took lowest")
    needs_review = bool(note_bits) or basic < 1000
    if basic < 1000:
        note_bits.append("unusually low basic — confirm")
    return {"employment_type": "full_time", "basic": basic,
            "hourly_rate": Decimal("0"), "needs_review": needs_review,
            "note": _joined(text, note_bits)}


def _joined(raw_text: str, bits: list[str]) -> str:
    base = f'Expected salary text: "{raw_text}"'
    return base + ("; " + "; ".join(bits) if bits else "")


def _clean(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _parse_dob(value) -> tuple[dt.date | None, str | None]:
    if isinstance(value, dt.datetime):
        d = value.date()
    elif isinstance(value, dt.date):
        d = value
    else:
        return None, None
    today = dt.date.today()
    if d > today:
        return d, "DOB is in the future — likely a data-entry error, confirm"
    if today.year - d.year > 80 or today.year - d.year < 14:
        return d, "DOB gives an unusual age — confirm"
    return d, None


def _is_foreign(bank: str | None) -> bool:
    b = (bank or "").lower()
    return "merchantrade" in b


def import_employees(db: Session, xlsx_path: str, sheet: str = "Form Responses 1") -> dict:
    """Import all registration rows as employees. Returns summary counts."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]

    start = db.scalar(select(func.count()).select_from(Employee)) or 0
    seq = start
    imported = 0
    flagged = 0
    for r in range(2, ws.max_row + 1):
        name = _clean(ws.cell(r, COL["name"]).value)
        if not name:
            continue
        seq += 1
        emp_code = f"ONNI{seq:04d}"
        sal = parse_expected_salary(ws.cell(r, COL["expected"]).value,
                                    _clean(ws.cell(r, COL["position"]).value) or "")
        dob, dob_note = _parse_dob(ws.cell(r, COL["dob"]).value)
        bank = _clean(ws.cell(r, COL["bank"]).value)
        foreign = _is_foreign(bank)
        note = sal["note"] + (f"; {dob_note}" if dob_note else "")
        needs_review = sal["needs_review"] or bool(dob_note)

        emp = Employee(
            emp_code=emp_code, reg_ref=_clean(ws.cell(r, COL["reg_id"]).value),
            name=name, email=_clean(ws.cell(r, COL["email"]).value),
            phone=_clean(ws.cell(r, COL["phone"]).value),
            nric=_clean(ws.cell(r, COL["nric"]).value), dob=dob,
            bank_name=bank, bank_account=_clean(ws.cell(r, COL["account"]).value),
            position=_clean(ws.cell(r, COL["position"]).value),
            employment_type=sal["employment_type"], basic_salary=sal["basic"],
            hourly_rate=sal["hourly_rate"],
            epf_enabled=(sal["employment_type"] == "full_time" and not foreign),
            socso_enabled=True, is_foreign=foreign,
            active=True, needs_review=needs_review, import_note=note,
        )
        db.add(emp)
        imported += 1
        flagged += int(needs_review)
    db.commit()
    return {"imported": imported, "flagged": flagged}
